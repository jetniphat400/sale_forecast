"""Reads product rows from reference/pricelist.xlsx, visible sheets only."""
import logging

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

_HEADER_ROW = 3  # 1-indexed row in the workbook holding column names
_DATA_START_ROW = 5  # 1-indexed row where product data begins

_REQUIRED_HEADERS = {
    "code": "Product Code",
    "category": "Product Cate.",
    "type": "Product Type",
    "description": "Product Description",
    "business": "Business",
}


def _find_col(headers: list, keyword: str):
    for idx, h in enumerate(headers):
        if h and keyword.lower() in str(h).lower():
            return idx
    return None


def load_visible_product_rows(path: str) -> pd.DataFrame:
    """Load product rows from every visible sheet in the pricelist workbook.

    Returns a DataFrame with columns: sheet, business, category, type,
    description, code. Logs which sheets were included/skipped and how
    many rows were read per sheet.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_state != "visible":
            logger.info("Skipping sheet %r: sheet_state=%r (not visible)", name, ws.sheet_state)
            continue

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        if len(rows) < _HEADER_ROW:
            logger.info("Skipping sheet %r: fewer than %d rows", name, _HEADER_ROW)
            continue

        header = rows[_HEADER_ROW - 1]
        code_idx = _find_col(header, _REQUIRED_HEADERS["code"])
        cate_idx = _find_col(header, _REQUIRED_HEADERS["category"])
        type_idx = _find_col(header, _REQUIRED_HEADERS["type"])
        desc_idx = _find_col(header, _REQUIRED_HEADERS["description"])
        biz_idx = _find_col(header, _REQUIRED_HEADERS["business"])

        if code_idx is None:
            logger.info("Skipping sheet %r: visible but no 'Product Code' column found", name)
            continue

        data_rows = rows[_DATA_START_ROW - 1:]
        kept, dropped_no_code = 0, 0
        for r in data_rows:
            code = r[code_idx] if code_idx < len(r) else None
            if code is None or str(code).strip() == "":
                dropped_no_code += 1
                continue
            records.append({
                "sheet": name,
                "business": r[biz_idx] if biz_idx is not None and biz_idx < len(r) else None,
                "category": r[cate_idx] if cate_idx is not None and cate_idx < len(r) else None,
                "type": r[type_idx] if type_idx is not None and type_idx < len(r) else None,
                "description": r[desc_idx] if desc_idx is not None and desc_idx < len(r) else None,
                "code": str(code).strip(),
            })
            kept += 1
        logger.info(
            "Sheet %r: included, %d product rows kept, %d rows dropped (no code)",
            name, kept, dropped_no_code,
        )

    df = pd.DataFrame.from_records(records)
    logger.info("Total product rows loaded from visible sheets: %d", len(df))
    return df
