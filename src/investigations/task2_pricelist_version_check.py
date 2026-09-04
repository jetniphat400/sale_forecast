"""Task 2 -- check which pricelist sheet (hidden Version1 vs visible Version2/current)
each of the no-history/no-sale items appears in. Method and limits per the task brief:
this can only show "already existed in a prior version" vs "new to the current version",
never a calendar date, since the workbook carries no date field for when a row was added.
"""
import os
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # src/ (moved to src/investigations/)
from pricelist_reader import _find_col, _HEADER_ROW, _DATA_START_ROW, _REQUIRED_HEADERS

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SUMMARY_DIR = os.path.join(PROJECT_ROOT, "output", "summary")
PATH = os.path.join(PROJECT_ROOT, "reference", "pricelist.xlsx")


def read_all_sheets(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        if len(rows) < _HEADER_ROW:
            print(f"Skipping sheet {name!r}: fewer than {_HEADER_ROW} rows")
            continue
        header = rows[_HEADER_ROW - 1]
        code_idx = _find_col(header, _REQUIRED_HEADERS["code"])
        if code_idx is None:
            print(f"Skipping sheet {name!r}: no 'Product Code' column found (state={ws.sheet_state})")
            continue
        data_rows = rows[_DATA_START_ROW - 1:]
        for r in data_rows:
            code = r[code_idx] if code_idx < len(r) else None
            if code is None or str(code).strip() == "":
                continue
            records.append({"sheet": name, "sheet_state": ws.sheet_state, "code": str(code).strip()})
    return pd.DataFrame.from_records(records)


if __name__ == "__main__":
    df = read_all_sheets(PATH)
    df.to_csv(os.path.join(SUMMARY_DIR, "task2_pricelist_all_sheets_all_codes.csv"), index=False)
    print(f"Total rows across ALL sheets (visible+hidden): {len(df)}")
    print(df.groupby(["sheet", "sheet_state"]).size())

    target16 = ['EEE-F-FL-1040030100', 'EEE-F-FL-5920-353-01100', 'EEE-F-FL-5920-353-01600',
                'EEE-F-FL-5920-353-02600', 'EEE-F-FL-5920-353-06600', 'FC-A-38-00203',
                'HS-F-99-0181', 'HS-F-99-1151', 'HS-F-99-1181', 'HS-F-99-1211H22',
                'HS-F-99-1241H03', 'HS-F-99-2091N', 'HS-F-99-3031', 'HS-F-99-3121',
                'HS-F-99-3331', 'HS-F-99-3361']
    sub = df[df["code"].isin(target16)]
    sub.to_csv(os.path.join(SUMMARY_DIR, "task2_pricelist_version_evidence_16items.csv"), index=False)
    print("\nPer-item sheet membership (across ALL sheets, visible+hidden):")
    for code in target16:
        rows = sub[sub["code"] == code]
        sheets = list(zip(rows["sheet"], rows["sheet_state"]))
        print(f"  {code}: {sheets if sheets else 'NOT FOUND IN ANY SHEET (visible or hidden)'}")
